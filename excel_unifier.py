#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel Unifier - 통일되지 않은 엑셀 파일들을 분석하고 통합하는 도구
"""

import pandas as pd
import os
from pathlib import Path
from typing import List, Dict, Tuple, Optional
from fuzzywuzzy import fuzz
from collections import defaultdict
import json
from dotenv import load_dotenv

# .env 파일 로드
load_dotenv()

# KFTA 파서 import
try:
    from kfta_parser import KFTAParser
except ImportError:
    KFTAParser = None


class ExcelUnifier:
    def __init__(
        self,
        similarity_threshold: int = 85,
        use_ai: bool = False,
        gemini_api_key: Optional[str] = None
    ):
        """
        엑셀 통합기 초기화

        Args:
            similarity_threshold: 유사도 임계값 (0-100, 기본값 85)
            use_ai: AI 기반 매칭 사용 여부 (기본값 False)
            gemini_api_key: Gemini API 키 (없으면 환경변수에서 읽음)
        """
        self.similarity_threshold = similarity_threshold
        self.use_ai = use_ai
        self.dataframes = []
        self.column_mappings = {}
        self.unified_columns = []

        # AI 모드 초기화
        self.ai_matcher = None
        if use_ai:
            try:
                from ai_matcher import GeminiMatcher
                self.ai_matcher = GeminiMatcher(api_key=gemini_api_key)
                print("🤖 AI 모드 활성화 (Gemini API)")
            except ImportError:
                print("⚠️  ai_matcher 모듈을 찾을 수 없습니다. 기본 모드로 전환합니다.")
                self.use_ai = False
            except Exception as e:
                print(f"⚠️  AI 모드 초기화 실패: {str(e)}. 기본 모드로 전환합니다.")
                self.use_ai = False

    def load_excel_files(self, file_paths: List[str]) -> None:
        """여러 엑셀 파일 로드 (모든 시트 포함)"""
        print(f"📂 {len(file_paths)}개의 파일을 로드합니다...")

        for file_path in file_paths:
            try:
                # 엑셀 파일 읽기 (.xlsx, .xls 모두 지원)
                if file_path.endswith('.csv'):
                    df = pd.read_csv(file_path)
                    self.dataframes.append({
                        'path': file_path,
                        'sheet': None,
                        'data': df,
                        'columns': list(df.columns)
                    })
                    print(f"  ✓ {os.path.basename(file_path)}: {len(df)}행, {len(df.columns)}개 컬럼")
                else:
                    # 엑셀 파일의 모든 시트 읽기
                    excel_file = pd.ExcelFile(file_path)
                    sheet_names = excel_file.sheet_names

                    print(f"  📄 {os.path.basename(file_path)}: {len(sheet_names)}개 시트 발견")

                    for sheet_name in sheet_names:
                        df = pd.read_excel(file_path, sheet_name=sheet_name)

                        # 빈 시트 건너뛰기
                        if df.empty or len(df.columns) == 0:
                            print(f"    ⊘ 시트 '{sheet_name}': 빈 시트 (건너뜀)")
                            continue

                        self.dataframes.append({
                            'path': file_path,
                            'sheet': sheet_name,
                            'data': df,
                            'columns': list(df.columns)
                        })
                        print(f"    ✓ 시트 '{sheet_name}': {len(df)}행, {len(df.columns)}개 컬럼")

            except Exception as e:
                print(f"  ✗ {file_path} 로드 실패: {str(e)}")

    def analyze_columns(self) -> Dict[str, List[str]]:
        """
        모든 파일의 컬럼명을 분석하고 유사한 컬럼끼리 그룹화

        Returns:
            컬럼 그룹 딕셔너리 {통합컬럼명: [원본컬럼명들]}
        """
        print("\n🔍 컬럼명 분석 중...")

        # 키워드 기반 컬럼 매핑 규칙
        # 순서 중요: 더 구체적인 것을 먼저 배치
        keyword_mappings = {
            # 강원교총 표준 매핑 (우선순위 높음)
            '현재교육청': ['현재교육청', '현 교육청', '소속교육청', '원교육청', '현재 교육청'],
            '현재본청': ['현재본청', '현재 본청', '현재학교', '현 본청', '현재 학교', '소속학교', '소속본청', '현재분회', '현재 분회'],
            '대응': ['대응', '대 응', '이름', '성명'],
            '발령교육청': ['발령교육청', '발령 교육청', '전입교육청', '배치교육청', '발령교 육청'],
            '발령본청': ['발령본청', '발령 본청', '전입학교', '배치학교', '발령학교', '발령 학교', '전보학교', '발령분회', '발령 분회'],
            '과목': ['과목', '교과', '담당과목', '과 목', '교 과'],
            '직위': ['직위', '직 위', '보직', '직급'],
            '직종분류': ['직종분류', '직종 분류', '직종', '교원구분', '교사구분', '교직원구분'],
            '분류명': ['분류명', '분류 명', '분류'],
            '취급코드': ['취급코드', '취급 코드', '구분코드'],
            '시군구분': ['시군구분', '시군 구분', '시군', '지역구분'],
            '교호기호등': ['교호기호등', '교호 기호등', '교호기호', '고호기호등'],

            # 일반 매핑 (우선순위 낮음)
            '이름': ['이름', '성명', '학생명', '성함', '이 름', '교사명', '교원명'],
            '학교': ['학교', '대학교', '소속대학', '대학', '학 교', '본청'],
            '전공': ['전공', '전공분야', '전공과목', '학과', '전 공'],
            '학년': ['학년', '학 년', '연차'],
            '연락처': ['연락처', '전화번호', '휴대폰', '전화', '연 락 처', '핸드폰', 'HP', '휴대전화'],
            '이메일': ['이메일', '메일', 'email', 'e-mail', '이 메 일'],
            '주소': ['주소', '주 소', '거주지', '집주소'],
        }

        # 모든 컬럼명 수집
        all_columns = []
        for df_info in self.dataframes:
            all_columns.extend(df_info['columns'])

        # 중복 제거하되 빈도수 유지
        column_freq = defaultdict(int)
        for col in all_columns:
            column_freq[col] += 1

        unique_columns = list(column_freq.keys())

        # 발견된 모든 컬럼 출력
        print(f"  📋 발견된 컬럼: {unique_columns}")

        # 컬럼 그룹화
        column_groups = {}
        processed = set()

        # 먼저 키워드 기반 매핑 적용
        for unified_name, keywords in keyword_mappings.items():
            matched_cols = []
            for col in unique_columns:
                if col in processed:
                    continue
                col_normalized = col.strip().replace(' ', '').lower()
                for keyword in keywords:
                    keyword_normalized = keyword.strip().replace(' ', '').lower()
                    # 정확히 일치하거나 포함하는 경우
                    if col_normalized == keyword_normalized or keyword_normalized in col_normalized:
                        matched_cols.append(col)
                        processed.add(col)
                        print(f"  ✓ '{col}' → '{unified_name}' (키워드: '{keyword}')")
                        break

            if matched_cols:
                column_groups[unified_name] = matched_cols
                if len(matched_cols) > 1:
                    print(f"  📌 '{unified_name}' ← {matched_cols}")

        # 나머지 컬럼들은 유사도 기반으로 매핑
        for i, col1 in enumerate(unique_columns):
            if col1 in processed:
                continue

            # 현재 컬럼과 유사한 컬럼들 찾기
            similar_cols = [col1]
            processed.add(col1)

            for col2 in unique_columns[i+1:]:
                if col2 in processed:
                    continue

                # 유사도 계산 (AI 모드 또는 기본 모드)
                is_similar = False

                if self.use_ai and self.ai_matcher:
                    # AI 기반 유사도 계산
                    try:
                        result = self.ai_matcher.calculate_semantic_similarity(
                            col1, col2,
                            context="엑셀 컬럼명"
                        )
                        is_similar = result['is_similar']
                        if is_similar:
                            print(f"  🤖 AI 매칭: '{col1}' ↔ '{col2}' ({result['similarity']}%, {result['reason']})")
                    except Exception as e:
                        print(f"  ⚠️  AI 분석 실패, 기본 모드로 전환: {str(e)}")
                        # 실패 시 기본 모드로 fallback
                        similarity = fuzz.ratio(col1, col2)
                        is_similar = similarity >= self.similarity_threshold
                else:
                    # 기본 모드: Levenshtein Distance
                    similarity = fuzz.ratio(col1, col2)
                    is_similar = similarity >= self.similarity_threshold

                if is_similar:
                    similar_cols.append(col2)
                    processed.add(col2)

            # 가장 빈도가 높은 컬럼명을 대표 컬럼명으로 선택
            representative = max(similar_cols, key=lambda x: column_freq[x])
            column_groups[representative] = similar_cols

            if len(similar_cols) > 1:
                print(f"  📌 '{representative}' ← {similar_cols}")

        self.column_mappings = column_groups
        self.unified_columns = list(column_groups.keys())

        # 매핑되지 않은 컬럼 표시
        unmatched = [col for col in unique_columns if col not in processed]
        if unmatched:
            print(f"  ⚠️  매핑되지 않은 컬럼 ({len(unmatched)}개): {unmatched}")

        print(f"\n✅ 총 {len(column_groups)}개의 통합 컬럼 생성")

        return column_groups

    def normalize_value(self, value: str, value_type: str = 'general') -> str:
        """
        값 정규화

        Args:
            value: 정규화할 값
            value_type: 값의 타입 (general, school, name 등)
        """
        if pd.isna(value):
            return ""

        value = str(value).strip()

        # 학교명 정규화
        if value_type == 'school':
            # 대학교, 대학, 고등학교, 중학교 등 통일
            replacements = [
                ('大學校', ''),
                ('大学校', ''),
                ('大学', ''),
                ('대학교', ''),
                ('대학', ''),
                ('고등학교', ''),
                ('고교', ''),
                ('중학교', ''),
                ('중학', ''),
                (' ', ''),
            ]
            for old, new in replacements:
                value = value.replace(old, new)

        return value.lower()

    def find_similar_values(self, values: List[str], threshold: int = None) -> Dict[str, List[str]]:
        """
        유사한 값들을 그룹화

        Args:
            values: 값 리스트
            threshold: 유사도 임계값

        Returns:
            {대표값: [유사한 값들]} 딕셔너리
        """
        if threshold is None:
            threshold = self.similarity_threshold

        unique_values = list(set([str(v) for v in values if pd.notna(v) and str(v).strip()]))

        value_groups = {}
        processed = set()

        for i, val1 in enumerate(unique_values):
            if val1 in processed:
                continue

            similar_vals = [val1]
            processed.add(val1)

            for val2 in unique_values[i+1:]:
                if val2 in processed:
                    continue

                similarity = fuzz.ratio(val1, val2)
                if similarity >= threshold:
                    similar_vals.append(val2)
                    processed.add(val2)

            # 가장 긴 값을 대표값으로 (보통 더 완전한 형태)
            representative = max(similar_vals, key=len)
            value_groups[representative] = similar_vals

        return value_groups

    def unify_dataframes(self, key_columns: List[str] = None, output_format: str = 'auto') -> pd.DataFrame:
        """
        모든 데이터프레임을 통합

        Args:
            key_columns: 중복 판단에 사용할 키 컬럼들 (예: ['이름', '학교'])
            output_format: 출력 형식 ('auto', 'standard', 'kfta')
                - 'auto': 자동 감지 (기본값)
                - 'standard': 모든 컬럼 포함
                - 'kfta': 강원교총 표준 형식 (12개 컬럼)

        Returns:
            통합된 데이터프레임
        """
        print("\n🔄 데이터 통합 중...")

        if not self.column_mappings:
            self.analyze_columns()

        # 통합 데이터프레임 초기화
        unified_data = []

        # 각 파일의 데이터를 통일된 컬럼명으로 변환
        for df_info in self.dataframes:
            df = df_info['data'].copy()

            # KFTA 형식이고 파서가 사용 가능하면 특수 파싱 적용
            if output_format == 'kfta' and KFTAParser is not None:
                parser = KFTAParser(use_ai=self.use_ai, ai_matcher=self.ai_matcher)
                df_unified = parser.parse_dataframe(df)

                file_name = os.path.basename(df_info['path'])
                sheet_info = f" (시트: {df_info['sheet']})" if df_info.get('sheet') else ""
                print(f"  ✓ {file_name}{sheet_info}: {len(df_unified)}행 변환 (KFTA 파서 사용)")

                unified_data.append(df_unified)
                continue

            # KFTA 형식이지만 파서 미사용 시, 또는 일반 매핑 방식
            # 컬럼명 매핑
            rename_dict = {}
            for unified_col, original_cols in self.column_mappings.items():
                for orig_col in original_cols:
                    if orig_col in df.columns:
                        rename_dict[orig_col] = unified_col

            df_renamed = df.rename(columns=rename_dict)

            # 중복된 컬럼명 처리 (같은 이름의 컬럼이 여러 개 있는 경우)
            if df_renamed.columns.duplicated().any():
                # 중복 컬럼 제거 (첫 번째만 유지)
                df_renamed = df_renamed.loc[:, ~df_renamed.columns.duplicated()]

            # 누락된 컬럼 추가 (빈 값으로)
            for col in self.unified_columns:
                if col not in df_renamed.columns:
                    df_renamed[col] = ""

            # 통일된 컬럼만 선택
            df_unified = df_renamed[self.unified_columns]
            unified_data.append(df_unified)

            # 시트 정보 포함하여 출력
            file_name = os.path.basename(df_info['path'])
            sheet_info = f" (시트: {df_info['sheet']})" if df_info.get('sheet') else ""
            print(f"  ✓ {file_name}{sheet_info}: {len(df_unified)}행 변환")

        # 모든 데이터 결합
        result_df = pd.concat(unified_data, ignore_index=True)
        print(f"\n📊 통합 완료: 총 {len(result_df)}행")

        # 출력 형식 적용 (KFTA 파서를 사용한 경우 이미 표준 형식이므로 건너뜀)
        if output_format == 'kfta' and KFTAParser is None:
            # KFTA 파서를 사용하지 않은 경우에만 적용
            result_df = self._apply_kfta_format(result_df)
        elif output_format == 'auto':
            # 강원교총 형식 컬럼이 존재하는지 확인
            kfta_columns = ['현재교육청', '현재본청', '대응', '발령교육청', '발령본청']
            if any(col in result_df.columns for col in kfta_columns):
                # 이미 KFTA 컬럼이 있으면 포맷만 적용
                if len(result_df.columns) != 12:  # 12개 표준 컬럼이 아니면
                    result_df = self._apply_kfta_format(result_df)

        # 키 컬럼이 지정된 경우 중복 제거
        if key_columns:
            print(f"\n🔑 키 컬럼 {key_columns}로 중복 확인 중...")

            # 키 컬럼이 실제로 존재하는지 확인
            valid_keys = [k for k in key_columns if k in result_df.columns]

            if valid_keys:
                before_count = len(result_df)
                result_df = self._remove_duplicates_smart(result_df, valid_keys)
                after_count = len(result_df)
                removed = before_count - after_count

                if removed > 0:
                    print(f"  ✓ {removed}개의 중복 행 제거됨")
                else:
                    print(f"  ℹ 중복 행이 발견되지 않음")

        return result_df

    def _remove_duplicates_smart(self, df: pd.DataFrame, key_columns: List[str]) -> pd.DataFrame:
        """
        스마트 중복 제거 - 유사한 값도 같은 것으로 간주
        """
        # 정규화된 키 컬럼 생성
        normalized_df = df.copy()

        for col in key_columns:
            if col in df.columns:
                # 값 정규화
                col_type = 'school' if '학교' in col else 'general'
                normalized_df[f'{col}_normalized'] = df[col].apply(
                    lambda x: self.normalize_value(x, col_type)
                )

        # 정규화된 값으로 중복 제거
        normalized_keys = [f'{col}_normalized' for col in key_columns if col in df.columns]

        if normalized_keys:
            # 중복 중 첫 번째 행 유지 (가장 완전한 데이터를 가진 행 선택)
            result_df = normalized_df.drop_duplicates(subset=normalized_keys, keep='first')

            # 정규화 컬럼 제거
            result_df = result_df.drop(columns=normalized_keys)
        else:
            result_df = df

        return result_df

    def _apply_kfta_format(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        강원교총 표준 형식으로 변환

        표준 컬럼 순서:
        1. 현재교육청
        2. 현재본청
        3. 대응
        4. 발령교육청
        5. 발령본청
        6. 과목
        7. 직위
        8. 직종분류
        9. 분류명
        10. 취급코드
        11. 시군구분
        12. 교호기호등
        """
        print("\n📋 강원교총 표준 형식 적용 중...")

        # 표준 컬럼 순서
        KFTA_COLUMNS = [
            '현재교육청',
            '현재본청',
            '대응',
            '발령교육청',
            '발령본청',
            '과목',
            '직위',
            '직종분류',
            '분류명',
            '취급코드',
            '시군구분',
            '교호기호등'
        ]

        result_df = df.copy()

        # 누락된 컬럼 추가 (빈 값으로)
        for col in KFTA_COLUMNS:
            if col not in result_df.columns:
                result_df[col] = ""
                print(f"  ℹ 컬럼 '{col}' 추가 (빈 값)")

        # 표준 순서대로 컬럼 재정렬
        result_df = result_df[KFTA_COLUMNS]

        print(f"  ✓ 표준 형식 적용 완료: {len(KFTA_COLUMNS)}개 컬럼")

        return result_df

    def save_unified_excel(self, output_path: str, df: pd.DataFrame = None) -> None:
        """통합된 데이터를 엑셀 파일로 저장 (타시도 학교 빨간색 표시)"""
        if df is None:
            df = self.unify_dataframes()

        print(f"\n💾 결과 저장 중: {output_path}")

        # 엑셀로 저장
        df.to_excel(output_path, index=False, engine='openpyxl')

        # 타시도 학교 셀 색상 표시
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill

            wb = load_workbook(output_path)
            ws = wb.active

            # 빨간색 배경 설정
            red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

            # 타시도 지역 리스트
            OTHER_REGIONS = [
                '서울', '부산', '대구', '인천', '광주', '대전', '울산', '세종',
                '경기', '충북', '충남', '전북', '전남', '경북', '경남', '제주'
            ]

            # 현재교육청과 발령교육청, 현재본청, 발령본청 컬럼 인덱스 찾기
            columns = list(df.columns)
            col_indices = {}
            for col_name in ['현재교육청', '발령교육청', '현재본청', '발령본청']:
                if col_name in columns:
                    # Excel 컬럼은 1부터 시작 (A=1, B=2, ...)
                    col_indices[col_name] = columns.index(col_name) + 1

            # 각 행 검사 (헤더 제외, 데이터는 2행부터)
            colored_count = 0
            for row_idx in range(2, len(df) + 2):  # Excel은 1-based, 헤더가 1행
                is_other_region = False

                # 교육청이 빈 문자열인지 확인
                if '현재교육청' in col_indices or '발령교육청' in col_indices:
                    for col_name in ['현재교육청', '발령교육청']:
                        if col_name in col_indices:
                            cell_value = ws.cell(row=row_idx, column=col_indices[col_name]).value
                            # 교육청이 비어있고 학교명이 있으면 타시도일 가능성
                            if not cell_value or str(cell_value).strip() == '':
                                # 해당 본청 확인
                                본청_col = '현재본청' if col_name == '현재교육청' else '발령본청'
                                if 본청_col in col_indices:
                                    school_value = ws.cell(row=row_idx, column=col_indices[본청_col]).value
                                    if school_value and str(school_value).strip():
                                        is_other_region = True
                                        break

                # 학교명에 타시도 지역명이 포함되어 있는지 확인
                if not is_other_region:
                    for col_name in ['현재본청', '발령본청']:
                        if col_name in col_indices:
                            cell_value = ws.cell(row=row_idx, column=col_indices[col_name]).value
                            if cell_value:
                                for region in OTHER_REGIONS:
                                    if region in str(cell_value):
                                        is_other_region = True
                                        break
                            if is_other_region:
                                break

                # 타시도 학교면 해당 행의 학교명 셀들을 빨간색으로
                if is_other_region:
                    for col_name in ['현재본청', '발령본청']:
                        if col_name in col_indices:
                            ws.cell(row=row_idx, column=col_indices[col_name]).fill = red_fill
                    colored_count += 1

            wb.save(output_path)

            if colored_count > 0:
                print(f"  🔴 타시도 학교 {colored_count}개 빨간색 표시 완료")

        except Exception as e:
            print(f"  ⚠️  셀 색상 표시 실패: {str(e)} (데이터는 정상 저장됨)")

        print(f"  ✓ 저장 완료: {len(df)}행, {len(df.columns)}개 컬럼")

    def generate_report(self, output_path: str = None) -> str:
        """분석 리포트 생성"""
        report = []
        report.append("=" * 60)
        report.append("📋 Excel Unifier 분석 리포트")
        report.append("=" * 60)
        report.append(f"\n총 파일 수: {len(self.dataframes)}")

        total_rows = sum(len(df_info['data']) for df_info in self.dataframes)
        report.append(f"총 행 수: {total_rows}")

        report.append(f"\n컬럼 매핑:")
        for unified_col, original_cols in self.column_mappings.items():
            if len(original_cols) > 1:
                report.append(f"  • {unified_col}")
                for orig in original_cols:
                    if orig != unified_col:
                        report.append(f"    - {orig}")

        report_text = "\n".join(report)

        if output_path:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(report_text)

        return report_text


def main():
    """메인 실행 함수"""
    import argparse

    parser = argparse.ArgumentParser(
        description='여러 엑셀 파일을 분석하고 통일된 양식으로 통합합니다.'
    )
    parser.add_argument(
        'files',
        nargs='+',
        help='통합할 엑셀 파일 경로들'
    )
    parser.add_argument(
        '-o', '--output',
        default='unified_output.xlsx',
        help='출력 파일명 (기본값: unified_output.xlsx)'
    )
    parser.add_argument(
        '-k', '--key-columns',
        nargs='+',
        help='중복 판단에 사용할 키 컬럼명들 (예: 이름 학교)'
    )
    parser.add_argument(
        '-t', '--threshold',
        type=int,
        default=85,
        help='유사도 임계값 0-100 (기본값: 85)'
    )
    parser.add_argument(
        '-r', '--report',
        help='분석 리포트 저장 경로'
    )
    parser.add_argument(
        '--ai',
        action='store_true',
        help='AI 기반 매칭 사용 (Gemini API 필요)'
    )
    parser.add_argument(
        '--api-key',
        help='Gemini API 키 (없으면 환경변수 GEMINI_API_KEY 사용)'
    )

    args = parser.parse_args()

    # ExcelUnifier 실행
    unifier = ExcelUnifier(
        similarity_threshold=args.threshold,
        use_ai=args.ai,
        gemini_api_key=args.api_key
    )
    unifier.load_excel_files(args.files)
    unifier.analyze_columns()

    # 데이터 통합
    unified_df = unifier.unify_dataframes(key_columns=args.key_columns)

    # 결과 저장
    unifier.save_unified_excel(args.output, unified_df)

    # 리포트 생성
    report = unifier.generate_report(args.report)
    print("\n" + report)

    print(f"\n✨ 완료! 결과 파일: {args.output}")


if __name__ == '__main__':
    main()
